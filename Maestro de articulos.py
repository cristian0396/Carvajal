import csv, openpyxl, getpass
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
from os import path
import os, errno, sys, re
import math
import time
import random
import threading
import datetime

if sys.version_info[0] == 2:
    from Tkinter import *
else:
    from tkinter import *

raiz = "C:\Maestro"
try:
    if os.path.exists("C:\Maestro"):
        print("Ya existe")
    else:
        os.mkdir(raiz)
except OSError as e:
    if e.errno != errno.EEXIST:
        raise
    print("Falló la creacion de carpeta" % raiz)

### RUTAS DEFINIDAS DE CARPETAS COMPARTIDAS ###
articulos_k40 = r'\\efiles\Temp\04_Carpak\Reportes_Oracle\CPKCO_ATRIBUTOS_ARTICULOS_K40.txt'
articulos_k42 = r'\\efiles\Temp\04_Carpak\Reportes_Oracle\CPKCO_ATRIBUTOS_ARTICULOS_K42.txt'
listas_k40 = r'\\efiles\Temp\04_Carpak\Reportes_Oracle\CPKCO_BOM_REPORTE_LISTAS_K40.txt'
listas_k42 = r'\\efiles\Temp\04_Carpak\Reportes_Oracle\CPKCO_BOM_REPORTE_LISTAS_K42.txt'
rutas_k40 = r'\\efiles\Temp\04_Carpak\Reportes_Oracle\CPKCO_BOM_REPORTE_RUTAS_K40.txt'
rutas_k42 = r'\\efiles\Temp\04_Carpak\Reportes_Oracle\CPKCO_BOM_REPORTE_RUTAS_K42.txt'

def convertir_articulos_k40(articulos_k40):
    """TOMA EL ARCHIVO ATRIBUTOS_ARTICULOS_K40 CVS DE LAS CARPETAS COMPARTIDAS Y LO CONVIERTE DE TEXTO A COLUMNAS"""
    f = open(articulos_k40)
    csv.register_dialect('colons', delimiter='|')
    reader = csv.reader(f, dialect='colons')
    wb = Workbook()
    dest_filename = r"C:\Maestro\Maestro articulos k40.xlsx"
    ws = wb.worksheets[0]
    ws.title = "Articulos"
    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):
            column_letter = get_column_letter((column_index + 1))
            ws['%s%s'%(column_letter, (row_index + 1))].value = cell
    wb.save(filename = dest_filename)
    
def convertir_articulos_k42(articulos_k42):
    """TOMA EL ARCHIVO ATRIBUTOS_ARTICULOS_K42 CVS DE LAS CARPETAS COMPARTIDAS Y LO CONVIERTE DE TEXTO A COLUMNAS"""
    f = open(articulos_k42)
    csv.register_dialect('colons', delimiter='|')
    reader = csv.reader(f, dialect='colons')
    wb = Workbook()
    dest_filename = r"C:\Maestro\Maestro articulos k42.xlsx"
    ws = wb.worksheets[0]
    ws.title = "Articulos"
    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):
            column_letter = get_column_letter((column_index + 1))
            ws['%s%s'%(column_letter, (row_index + 1))].value = cell
    wb.save(filename = dest_filename)
    
def obtener_cod_articulo_k40():
    """TOMA EL ARCHIVO CONVERTIDO ATRIBUTOS_ARTICULOS_K40 Y EN UN NUEVO ARCHIVO FILTRA LOS CODIGOS PADRES QUE ESTEN ACTIVOS Y HAI"""
    articulos = Workbook()
    articulo = articulos.worksheets[0]
    libro_art = load_workbook('C:\Maestro\Maestro articulos k40.xlsx')
    hoja_art = libro_art.worksheets[0]
    cont = 1
    colores =['AMARILLO','AMBAR','AZUL','BEIGE','BLANCO','CAFE','CASTAÑO','DORADO','GRIS','LILA','MARRON','NARANJA','NEGRO','OPAL','PERLADO','PLATA','ROJO','ROSADO','TRANSLUCIDO','TRASLUCIDO','TRANSPARENTE','VERDE','VINOTINTO','VIOLETA']
    articulo.append(["Codigo","Nombre","Cliente","Categoria de inventario","Linea","Mercado","Tipo Terminacion","Organizacion","Estado","Cod.Componente","Cod.Modelo","Tipo Modelo","Diametro","Largo","Peso Producion","Formulacion","Material","Color","Ancho","Calibre","Cod.Insumo","Corrugado Master","Und.Empque Maestro","Corrugado Secundario","Und.Empque Secundaria","Plegadiza","Und.Empque Plegadiza","Particion","Und.Empque Particion","Bolsa/Rollo Master","Unidad Empaque Bolsa Master","Bolsa/Rollo Secundaria","Unidad Empaque Bolsa Secundaria","EXT_1501","EXT_WEL90","EXT_1503","EXT_WEL120","TER_3710","TER_70K","TER_50K","TER_RDK54","TER_GABLER","IMP_560","IMP_560R","IMP_580R","ETI_A","ETI_B","TER_KARV_200","TER_KARV_400","EMP_RENCO","TER_AUTOV","TER_ILLIG_P","TER_RDKP72","TER_RDKP54","TER_ILLIG_G","TER_TAIWANESA","TRO_KRAUSSE","PRE_HERLAN","PRE_SHULER","LIN_HERLAN","LIN_MALL","LIN_SCHULER","ENG_A","ENG_B","INY_FROMAG","INY_ORIENTE","LLE_COMADIS","SEL_AUTOMATICA","HOR_A","LAC_SPRIGMAG","MOL_TERMICOS","TER_COMMODORE","PESO VENTAS","VOLUMEN EMPAQUE PPAL","UNIDAD VOLUMEN EMPAQUE PPAL","YIELD","CATEGORIA COSTOS PAC","CATEGORIA PRODUCTO","VOLUMEN ORACLE CVR(M3)","PESO ORACLE CVR(Kg)","INTRACLASE CAJA","PESO BRUTO(Kg)","RADICADO","TIPO DE MOLDE","MOLDE 1","MOLDE 2","MOLDE 3","YIELD","DESPERDICIO","OP STD","DESCRIPCION OP","OP STD PRODUCCION","RECURSO","INVERSO(RENDIMIENTO)","IDENTIFICADOR DE BODEGAS","Desperdicio Componente","Desperdicio Insumo","Desperdicio Corr Master","Desperdicio Corr sec","Desperdicio Bolsa master","Desperdicio Bolsa sec","Descripcion larga"])
    nombre = r"C:\Maestro\MaestroK40.xlsx"
    for fila in range(2,hoja_art.max_row):
        if "Active" == hoja_art.cell(row= fila, column=37).value or "HAI" == hoja_art.cell(row= fila, column=37).value:
            cont += 1
            articulo.cell(row=cont,column=1, value=hoja_art.cell(row=fila,column=2).value) #CODIGO PADRE
            articulo.cell(row=cont,column=2, value=hoja_art.cell(row=fila,column=3).value) #DESCRIPCION CORTA
            articulo.cell(row=cont,column=3, value=hoja_art.cell(row=fila,column=12).value)#CLIENTE
            articulo.cell(row=cont,column=4, value=hoja_art.cell(row=fila,column=24).value)#CATEGORIA DE INVENTARIO
            cadena = hoja_art.cell(row=fila,column=1).value
            org = re.split(r'\ ',cadena)
            articulo.cell(row=cont,column=8, value=org[0])#ORGANIZACION
            articulo.cell(row=cont,column=9, value=hoja_art.cell(row=fila,column=37).value)#ESTADO            
            articulo.cell(row=cont,column=11, value=hoja_art.cell(row=fila,column=10).value)#COD MODELO
            if (hoja_art.cell(row=fila,column=131).value)is not None:
                articulo.cell(row=cont,column=15, value=float(hoja_art.cell(row=fila,column=131).value)*1000)#PESO PRODUCCION
                articulo.cell(row=cont,column=73, value=float(hoja_art.cell(row=fila,column=131).value)*1000)#PESO PRODUCCION
            cadena = hoja_art.cell(row=fila,column=24).value            
            if (cadena)is not None:
                material = re.split(r'\.',cadena)
                if len(material)>=1:
                    articulo.cell(row=cont,column=17, value=material[1])#MATERIAL
            cadena = hoja_art.cell(row=fila,column=23).value
            separado = re.split(r'\.',cadena)
            if (len(separado))>=3:                
                articulo.cell(row=cont,column=5, value=separado[1]) #LINEA
                articulo.cell(row=cont,column=6, value=separado[0]) #MERCADO
                articulo.cell(row=cont,column=7, value=separado[3]) #TIPO TERMINACION
            for color in colores:
                if re.search(color, hoja_art.cell(row=fila,column=3).value)is not None:
                    articulo.cell(row=cont,column=18, value=color) #COLOR
                    break
            if re.match("SP LAM", hoja_art.cell(row= fila, column=3).value)is not None:
                ac = hoja_art.cell(row= fila, column=3).value
                ac = ac[ac.find("X")-5:ac.find("X")+5]
                ac = re.findall(r'\d+.\d+',ac)
                if ac == [] or len(ac)<2:
                    articulo.cell(row=cont,column=19, value="Validar descripcion") #ANCHO
                    articulo.cell(row=cont,column=20, value="Validar descripcion") #CALIBRE
                elif len(ac[0])>3:
                    articulo.cell(row=cont,column=19, value=ac[1]) #ANCHO
                    articulo.cell(row=cont,column=20, value=ac[0]) #CALIBRE
                else:
                    articulo.cell(row=cont,column=19, value=ac[1]) #ANCHO
                    articulo.cell(row=cont,column=20, value="Validar descripcion") #CALIBRE            
            articulo.cell(row=cont,column=75, value=hoja_art.cell(row=fila,column=132).value)#UND VOLUMEN EMPAQUE PPAL
            articulo.cell(row=cont,column=77, value=hoja_art.cell(row=fila,column=25).value)#CATEGORIA COSTOS PAC
            articulo.cell(row=cont,column=78, value=hoja_art.cell(row=fila,column=23).value)#CATEGORIA PRODUCTO 
            articulo.cell(row=cont,column=79, value=hoja_art.cell(row=fila,column=133).value)#VOLUMEN ORACLE CVR
            articulo.cell(row=cont,column=80, value=hoja_art.cell(row=fila,column=131).value)#PESO ORACLE CVR (Kg)
            articulo.cell(row=cont,column=81, value=hoja_art.cell(row=fila,column=20).value)#INTERCLASE CAJA
            articulo.cell(row=cont,column=82, value=hoja_art.cell(row=fila,column=21).value)#INTERCLASE CAJA
            articulo.cell(row=cont,column=102, value=hoja_art.cell(row=fila,column=38).value)#DESCRIPCION LARGA
    articulos.save(nombre)
    articulos.close()
    print("Se guardó correctamente")
    return articulos

def obtener_cod_articulo_k42():
    """TOMA EL ARCHIVO CONVERTIDO ATRIBUTOS_ARTICULOS_K42 Y EN UN NUEVO ARCHIVO FILTRA LOS CODIGOS PADRES QUE ESTEN ACTIVOS Y HAI"""
    articulos = Workbook()
    articulo = articulos.worksheets[0]
    libro_art = load_workbook('C:\Maestro\Maestro articulos k42.xlsx')
    hoja_art = libro_art.worksheets[0]
    cont = 1
    colores =['AMARILLO','AMBAR','AZUL','BEIGE','BLANCO','CAFE','CASTAÑO','DORADO','GRIS','LILA','MARRON','NARANJA','NEGRO','OPAL','PERLADO','PLATA','ROJO','ROSADO','TRANSLUCIDO','TRASLUCIDO','TRANSPARENTE','VERDE','VINOTINTO','VIOLETA']
    articulo.append(["Codigo","Nombre","Cliente","Categoria de inventario","Linea","Mercado","Tipo Terminacion","Organizacion","Estado","Cod.Componente","Cod.Modelo","Tipo Modelo","Diametro","Largo","Peso Producion","Formulacion","Material","Color","Ancho","Calibre","Cod.Insumo","Corrugado Master","Und.Empque Maestro","Corrugado Secundario","Und.Empque Secundaria","Plegadiza","Und.Empque Plegadiza","Particion","Und.Empque Particion","Bolsa/Rollo Master","Unidad Empaque Bolsa Master","Bolsa/Rollo Secundaria","Unidad Empaque Bolsa Secundaria","EXT_1501","EXT_WEL90","EXT_1503","EXT_WEL120","TER_3710","TER_70K","TER_50K","TER_RDK54","TER_GABLER","IMP_560","IMP_560R","IMP_580R","ETI_A","ETI_B","TER_KARV_200","TER_KARV_400","EMP_RENCO","TER_AUTOV","TER_ILLIG_P","TER_RDKP72","TER_RDKP54","TER_ILLIG_G","TER_TAIWANESA","TRO_KRAUSSE","PRE_HERLAN","PRE_SHULER","LIN_HERLAN","LIN_MALL","LIN_SCHULER","ENG_A","ENG_B","INY_FROMAG","INY_ORIENTE","LLE_COMADIS","SEL_AUTOMATICA","HOR_A","LAC_SPRIGMAG","MOL_TERMICOS","TER_COMMODORE","PESO VENTAS","VOLUMEN EMPAQUE PPAL","UNIDAD VOLUMEN EMPAQUE PPAL","YIELD","CATEGORIA COSTOS PAC","CATEGORIA PRODUCTO","VOLUMEN ORACLE CVR(M3)","PESO ORACLE CVR(Kg)","INTRACLASE CAJA","PESO BRUTO(Kg)","RADICADO","TIPO DE MOLDE","MOLDE 1","MOLDE 2","MOLDE 3","YIELD","DESPERDICIO","OP STD","DESCRIPCION OP","OP STD PRODUCCION","RECURSO","INVERSO(RENDIMIENTO)","IDENTIFICADOR DE BODEGAS","Desperdicio Componente","Desperdicio Insumo","Desperdicio Corr Master","Desperdicio Corr sec","Desperdicio Bolsa master","Desperdicio Bolsa sec","Descripcion larga"])
    nombre = r"C:\Maestro\MaestroK42.xlsx"
    for fila in range(2,hoja_art.max_row):
        if "Active" == hoja_art.cell(row= fila, column=37).value or "HAI" == hoja_art.cell(row= fila, column=37).value:
            cont += 1
            articulo.cell(row=cont,column=1, value=hoja_art.cell(row=fila,column=2).value) #CODIGO PADRE
            articulo.cell(row=cont,column=2, value=hoja_art.cell(row=fila,column=3).value) #DESCRIPCION CORTA
            articulo.cell(row=cont,column=3, value=hoja_art.cell(row=fila,column=12).value)#CLIENTE
            articulo.cell(row=cont,column=4, value=hoja_art.cell(row=fila,column=24).value)#CATEGORIA DE INVENTARIO
            cadena = hoja_art.cell(row=fila,column=1).value
            org = re.split(r'\ ',cadena)
            articulo.cell(row=cont,column=8, value=org[0])#ORGANIZACION
            articulo.cell(row=cont,column=9, value=hoja_art.cell(row=fila,column=37).value)#ESTADO            
            articulo.cell(row=cont,column=11, value=hoja_art.cell(row=fila,column=10).value)#COD MODELO
            if (hoja_art.cell(row=fila,column=131).value)is not None:
                articulo.cell(row=cont,column=15, value=float(hoja_art.cell(row=fila,column=131).value)*1000)#PESO PRODUCCION
                articulo.cell(row=cont,column=73, value=float(hoja_art.cell(row=fila,column=131).value)*1000)#PESO PRODUCCION
            cadena = hoja_art.cell(row=fila,column=24).value            
            if (cadena)is not None:
                material = re.split(r'\.',cadena)
                if len(material)>=1:
                    articulo.cell(row=cont,column=17, value=material[1])#MATERIAL
            cadena = hoja_art.cell(row=fila,column=23).value
            separado = re.split(r'\.',cadena)
            if (len(separado))>=3:                
                articulo.cell(row=cont,column=5, value=separado[1]) #LINEA
                articulo.cell(row=cont,column=6, value=separado[0]) #MERCADO
                articulo.cell(row=cont,column=7, value=separado[3]) #TIPO TERMINACION
            for color in colores:
                if re.search(color, hoja_art.cell(row=fila,column=3).value)is not None:
                    articulo.cell(row=cont,column=18, value=color) #COLOR
                    break
            if re.match("SP LAM", hoja_art.cell(row= fila, column=3).value)is not None:
                ac = hoja_art.cell(row= fila, column=3).value
                ac = ac[ac.find("X")-5:ac.find("X")+5]
                ac = re.findall(r'\d+.\d+',ac)
                if ac == [] or len(ac)<2:
                    articulo.cell(row=cont,column=19, value="Validar descripcion") #ANCHO
                    articulo.cell(row=cont,column=20, value="Validar descripcion") #CALIBRE
                elif len(ac[0])>3:
                    articulo.cell(row=cont,column=19, value=ac[1]) #ANCHO
                    articulo.cell(row=cont,column=20, value=ac[0]) #CALIBRE
                else:
                    articulo.cell(row=cont,column=19, value=ac[1]) #ANCHO
                    articulo.cell(row=cont,column=20, value="Validar descripcion") #CALIBRE            
            articulo.cell(row=cont,column=75, value=hoja_art.cell(row=fila,column=132).value)#UND VOLUMEN EMPAQUE PPAL
            articulo.cell(row=cont,column=77, value=hoja_art.cell(row=fila,column=25).value)#CATEGORIA COSTOS PAC
            articulo.cell(row=cont,column=78, value=hoja_art.cell(row=fila,column=23).value)#CATEGORIA PRODUCTO 
            articulo.cell(row=cont,column=79, value=hoja_art.cell(row=fila,column=133).value)#VOLUMEN ORACLE CVR
            articulo.cell(row=cont,column=80, value=hoja_art.cell(row=fila,column=131).value)#PESO ORACLE CVR (Kg)
            articulo.cell(row=cont,column=81, value=hoja_art.cell(row=fila,column=20).value)#INTERCLASE CAJA
            articulo.cell(row=cont,column=82, value=hoja_art.cell(row=fila,column=21).value)#INTERCLASE CAJA
            articulo.cell(row=cont,column=102, value=hoja_art.cell(row=fila,column=38).value)#DESCRIPCION LARGA
    articulos.save(nombre)
    articulos.close()
    print("Se guardó correctamente")
    return articulos

def convertir_listas_k40(listas_k40):
    """TOMA EL ARCHIVO LISTAS_K40 CVS DE LAS CARPETAS COMPARTIDAS Y LO CONVIERTE DE TEXTO A COLUMNAS"""
    f = open(listas_k40)
    csv.register_dialect('colons', delimiter='|')
    reader = csv.reader(f, dialect='colons')
    wb = Workbook()
    dest_filename = r"C:\Maestro\Maestro Listas k40.xlsx"
    ws = wb.worksheets[0]
    ws.title = "Maestro de articulos"
    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):
            column_letter = get_column_letter((column_index + 1))
            ws['%s%s'%(column_letter, (row_index + 1))].value = cell
    wb.save(filename = dest_filename)
    print("Termina la conversion de articulos",time.asctime(time.localtime(time.time())))

def convertir_listas_k42(listas_k42):
    """TOMA EL ARCHIVO LISTAS_K42 CVS DE LAS CARPETAS COMPARTIDAS Y LO CONVIERTE DE TEXTO A COLUMNAS"""
    f = open(listas_k42)
    csv.register_dialect('colons', delimiter='|')
    reader = csv.reader(f, dialect='colons')
    wb = Workbook()
    dest_filename = r"C:\Maestro\Maestro Listas k42.xlsx"
    ws = wb.worksheets[0]
    ws.title = "Maestro de articulos"
    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):
            column_letter = get_column_letter((column_index + 1))
            ws['%s%s'%(column_letter, (row_index + 1))].value = cell
    wb.save(filename = dest_filename)
    
def obtener_corrugado_k40(cod_articulos):
    """TOMA EL ARCHIVO FILTRADO DE LOS ARTICULOS Y BUSCA EL CORRUGADO DE CADA ARTICULO, EN EL PROCESO SE OBTIENEN MUCHOS MAS DATOS DEL MISMO ARTICULO"""
    print("INICIA EL CORRUGADO",time.asctime(time.localtime(time.time())))
    hoja_corrugado = cod_articulos.worksheets[0]
    libro = load_workbook('C:\Maestro\Maestro Listas k40.xlsx')
    hoja = libro.worksheets[0]
    libro_rutas = load_workbook('C:\Maestro\Maestro rutas k40.xlsx')
    hoja_ruta = libro_rutas.worksheets[0]
    lista_SP =['SP BASE','SP CONTENEDOR','SP VASO','SP TUBO','MAQUILA','SP LAM','SP BANDEJA','SP CAZUELA','SP VISOR','SP ESTUCHE','SP COPA','SP SOBRECOPA','SP ETIQUETA','SP PLIEGO']
    lista_insumo =['ETIQUETA','FUNDA','FAJA','TARJETA','TAPA MAQUILA','MATMANUF']
    formulacion = ['R ', 'PASTILLA']
    colores =['AMARILLO','AMBAR','AZUL','BEIGE','BLANCO','CAFE','CASTAÑO','DORADO','GRIS','LILA','MARRON','NARANJA','NEGRO','OPAL','PERLADO','PLATA','ROJO','ROSADO','TRANSLUCIDO','TRASLUCIDO','TRANSPARENTE','VERDE','VINOTINTO','VIOLETA']
    corrugado = "CORRUGADO"
    nombre = r"C:\Maestro\MaestroK40.xlsx" 
    for articulo in range(1, hoja_corrugado.max_row):           #RECORRE LA HOJA CON TODOS LOS PADRES DE ARTICULOS
        codPadre = hoja_corrugado.cell(row= articulo, column=1).value
        for fila in range(2, hoja.max_row):                     #RECORRE LA HOJA DE LAS LISTAS BUSCANDO EL CODIGO PADRE PARA ENCONTRAR EL SP BASE
            if codPadre == hoja.cell(row= fila, column=2).value and (hoja.cell(row= fila, column=5).value)is None: #EVALUA QUE EL COD SEA IGUAL Y PRINCIPAL
                """Primero debe buscar si tiene corrugado, si lo tiene lo trae de una
                    sino, busca el SP BASE, SP CONTENEDOR, SP VASO, SP TUBO, MAQUILA: para obtener el corrugado"""
                if re.search("TAPA", hoja.cell(row=fila,column=3).value) is not None or re.search("TUBO", hoja.cell(row=fila,column=3).value) is not None:
                    if re.search("SP TAPA", hoja.cell(row=fila,column=8).value) is not None:
                        hoja_corrugado.cell(row=articulo,column=10, value=hoja.cell(row=fila,column=7).value)#CODIGO COMPONENTE
                for color in colores:
                    if re.search(color, hoja.cell(row=fila,column=3).value)is not None and (hoja_corrugado.cell(row=articulo,column=18).value)is None:
                        hoja_corrugado.cell(row=articulo,column=18, value=color) #COLOR
                        break
                    elif re.search(color, hoja.cell(row=fila,column=8).value)is not None and (hoja_corrugado.cell(row=articulo,column=18).value)is None:
                        hoja_corrugado.cell(row=articulo,column=18, value=color) #COLOR
                        break
                for formu in formulacion:
                    if re.match(formu, hoja.cell(row= fila, column=8).value)is not None:
                        hoja_corrugado.cell(row=articulo,column=16, value=hoja.cell(row= fila, column=7).value) #FORMULACION
                for insumo in lista_insumo:                
                    if re.match(insumo, hoja.cell(row= fila, column=8).value)is not None:
                        hoja_corrugado.cell(row=articulo,column=21, value=hoja.cell(row=fila,column=7).value)#CODIGO INSUMO
                        break
                if re.search("CORRUGADO", hoja.cell(row= fila, column=8).value)is not None or re.search("BOLSA",hoja.cell(row= fila, column=8).value)is not None:
                    if re.search("BOLSA",hoja.cell(row= fila, column=8).value)is not None:
                        if (hoja_corrugado.cell(row= articulo, column=30).value) is None:
                            hoja_corrugado.cell(row=articulo,column=30, value=hoja.cell(row= fila, column=7).value) #BOLSA/ROLLO MASTER
                            hoja_corrugado.cell(row=articulo,column=104, value=hoja.cell(row= fila, column=8).value) #DESCRIPCION BOLSA MASTER
                            hoja_corrugado.cell(row=articulo,column=99, value=(1-float(hoja.cell(row= fila, column=13).value))) #DESPERDICIO BOLSA MASTER
                            if float(hoja.cell(row= fila, column=12).value)*1 != 0 :
                                hoja_corrugado.cell(row=articulo,column=31, value=int(1/float(hoja.cell(row= fila, column=12).value))) #UND.EMPQUE BOLSA MASTER
                            else:
                                hoja_corrugado.cell(row=articulo,column=31, value=hoja.cell(row= fila, column=12).value) #UND.EMPQUE BOLSA MASTER
                        else:
                            bolsa1=hoja_corrugado.cell(row= articulo, column=104).value
                            bolsa2=hoja.cell(row= fila, column=8).value
                            bol1 = bolsa1[bolsa1.find("X")-5:bolsa1.find("X")+5]                            
                            bol2 = bolsa2[bolsa2.find("X")-5:bolsa2.find("X")+5]
                            bol1 = re.findall(r'\d+',bolsa1)
                            bol2 = re.findall(r'\d+',bolsa2)
                            if len(bol1) == 2 and len(bol2)==2:
                                if float(bol1[0])*float(bol1[1]) > float(bol2[0])*float(bol2[1]):
                                    hoja_corrugado.cell(row=articulo,column=32, value=hoja.cell(row= fila, column=7).value) #BOLSA ROLLO SEC
                                    hoja_corrugado.cell(row=articulo,column=101, value=(1-float(hoja.cell(row= fila, column=13).value))) #DESPERDICIO BOLSA SEC
                                    if float(hoja.cell(row= fila, column=12).value)*1 != 0 :
                                        hoja_corrugado.cell(row=articulo,column=33, value=int(1/float(hoja.cell(row= fila, column=12).value))) #UND.EMPQUE BOLSA SECUNDARIA
                                    else:
                                        hoja_corrugado.cell(row=articulo,column=33, value=hoja.cell(row= fila, column=12).value) #UND.EMPQUE BOLSA SECUNDARIA
                                else:
                                    hoja_corrugado.cell(row=articulo,column=32, value=hoja_corrugado.cell(row=articulo,column=30).value) #BOLSA ROLLO SEC
                                    hoja_corrugado.cell(row=articulo,column=33, value=hoja_corrugado.cell(row=articulo,column=31).value) #BOLSA ROLLO SEC
                                    hoja_corrugado.cell(row=articulo,column=101, value=hoja_corrugado.cell(row=articulo,column=99).value) #DESPERDICIO BOLSA SEC
                                    hoja_corrugado.cell(row=articulo,column=30, value=hoja.cell(row= fila, column=7).value) #BOLSA ROLLO SEC
                                    hoja_corrugado.cell(row=articulo,column=99, value=(1-float(hoja.cell(row= fila, column=13).value))) #DESPERDICIO BOLSA SEC
                                    if float(hoja.cell(row= fila, column=12).value)*1 != 0 :
                                        hoja_corrugado.cell(row=articulo,column=31, value=int(1/float(hoja.cell(row= fila, column=12).value))) #UND.EMPQUE BOLSA SECUNDARIA
                                    else:
                                        hoja_corrugado.cell(row=articulo,column=31, value=hoja.cell(row= fila, column=12).value) #UND.EMPQUE BOLSA SECUNDARIA
                            else:
                                hoja_corrugado.cell(row=articulo,column=32, value="Validar descripcion")
                                hoja_corrugado.cell(row=articulo,column=33, value="Validar descripcion")                        
                    else:
                        if (hoja_corrugado.cell(row=articulo,column=22).value)is None:
                            hoja_corrugado.cell(row=articulo,column=22, value=hoja.cell(row= fila, column=7).value) #CORRUGADO MASTER
                            hoja_corrugado.cell(row=articulo,column=103, value=hoja.cell(row= fila, column=8).value) #DESCRIPCION CORRUGADO MASTER
                            hoja_corrugado.cell(row=articulo,column=98, value=(1-float(hoja.cell(row= fila, column=13).value))) #DESPERDICIO CORRUGADO MASTER
                            if float(hoja.cell(row= fila, column=12).value)*1 != 0:
                                hoja_corrugado.cell(row=articulo,column=23, value=int(1/float(hoja.cell(row= fila, column=12).value))) #UND.EMPQUE MAESTRO
                            else:
                                hoja_corrugado.cell(row=articulo,column=23, value=hoja.cell(row= fila, column=12).value) #UND.EMPQUE MAESTRO
                        else:
                            cor1 = hoja_corrugado.cell(row=articulo,column=103).value
                            cor2 = hoja.cell(row= fila, column=8).value
                            tam1 = re.findall(r'\d+',cor1)
                            tam2 = re.findall(r'\d+',cor2)
                            if len(tam1)>3 and len(tam2)>3:
                                if float(tam1[1])*float(tam1[2])*float(tam1[3]) > float(tam2[1])*float(tam2[2])*float(tam2[3]):
                                    hoja_corrugado.cell(row=articulo,column=24, value=hoja.cell(row= fila, column=7).value) #CORRUGADO SECUNDARIO
                                    hoja_corrugado.cell(row=articulo,column=99, value=(1-float(hoja.cell(row= fila, column=13).value))) #DESPERDICIO CORRUGADO SEC
                                    if float(hoja.cell(row= fila, column=12).value)*1 != 0:
                                        hoja_corrugado.cell(row=articulo,column=25, value=int(1/float(hoja.cell(row= fila, column=12).value))) #UND.EMPQUE SECUNDARIO
                                    else:
                                        hoja_corrugado.cell(row=articulo,column=25, value=hoja.cell(row= fila, column=12).value) #UND.EMPQUE SECUNDARIO
                                else:
                                    hoja_corrugado.cell(row=articulo,column=24, value=hoja_corrugado.cell(row=articulo,column=22).value) #CORRUGADO SECUNDARIO
                                    hoja_corrugado.cell(row=articulo,column=25, value=hoja_corrugado.cell(row=articulo,column=23).value) #UND EMPQUE SECUNDARIO
                                    hoja_corrugado.cell(row=articulo,column=99, value=(1-float(hoja.cell(row= fila, column=13).value))) #DESPERDICIO CORRUGADO SEC
                                    hoja_corrugado.cell(row=articulo,column=22, value=hoja.cell(row= fila, column=7).value) #CORRUGADO MASTER
                                    hoja_corrugado.cell(row=articulo,column=98, value=(1-float(hoja.cell(row= fila, column=13).value))) #DESPERDICIO CORRUGADO MASTER
                                    if float(hoja.cell(row= fila, column=12).value)*1 != 0:
                                        hoja_corrugado.cell(row=articulo,column=23, value=int(1/float(hoja.cell(row= fila, column=12).value))) #UND.EMPQUE MAESTRO
                                    else:
                                        hoja_corrugado.cell(row=articulo,column=23, value=hoja.cell(row= fila, column=12).value) #UND.EMPQUE MAESTRO                
                else:
                    for sp in lista_SP:
                        if re.search(sp, hoja.cell(row= fila, column=8).value)is not None:
                            hoja_corrugado.cell(row=articulo,column=10, value=hoja.cell(row=fila,column=7).value)#CODIGO COMPONENTE SP
                            hoja_corrugado.cell(row=articulo,column=96, value=(1-float(hoja.cell(row= fila, column=13).value))) #DESPERDICIO COMPONENTE
                            codCorrugado = hoja.cell(row= fila, column=7).value
                            for componente in range(1, hoja.max_row):   #RECORRE LAS LISTAS BUSCANDO EL COD DEL SP PARA ENCONTRAR EL COMPONENTE
                                if codCorrugado == hoja.cell(row= componente, column=2).value and (hoja.cell(row= componente, column=5).value)is None:
                                    for color in colores:
                                        if re.search(color, hoja.cell(row=componente, column=3).value)is not None and  (hoja_corrugado.cell(row=articulo,column=18).value)is None:
                                            hoja_corrugado.cell(row=articulo,column=18, value=color) #COLOR
                                            break
                                        elif re.search(color, hoja.cell(row=componente,column=8).value)is not None and (hoja_corrugado.cell(row=articulo,column=18).value)is None:
                                            hoja_corrugado.cell(row=articulo,column=18, value=color) #COLOR
                                            break
                                    des = hoja.cell(row= componente, column=8).value
                                    for formu in formulacion:
                                        if re.match(formu, hoja.cell(row=componente, column=8).value)is not None and (hoja_corrugado.cell(row=articulo,column=16).value)is None:
                                            hoja_corrugado.cell(row=articulo,column=16, value=hoja.cell(row= componente, column=7).value) #FORMULACION                                    
                                    if re.search("BOLSA",hoja.cell(row= componente, column=8).value)is not None and (hoja_corrugado.cell(row= articulo, column=32).value) is None:
                                        hoja_corrugado.cell(row=articulo,column=32, value=hoja.cell(row= componente, column=7).value) #BOLSA/ROLLO MASTER
                                        hoja_corrugado.cell(row=articulo,column=33, value=int(1/float(hoja.cell(row= componente, column=12).value))) #UND EMPQUE BOLSA MASTER
                                        hoja_corrugado.cell(row=articulo,column=100, value=(1-float(hoja.cell(row= fila, column=13).value))) #DESPERDICIO BOLSA MASTER
                                    
                                    if re.search(corrugado, des)is not None:
                                        if (hoja_corrugado.cell(row=articulo,column=22).value)is None:
                                            hoja_corrugado.cell(row=articulo,column=22, value=hoja.cell(row= componente, column=7).value) #CORRUGADO MASTER
                                            hoja_corrugado.cell(row=articulo,column=103, value=hoja.cell(row= componente, column=8).value) #DESCRIPCION CORRUGADO MASTER
                                            hoja_corrugado.cell(row=articulo,column=98, value=(1-float(hoja.cell(row= componente, column=13).value))) #DESPERDICIO CORRUGADO MASTER
                                            if float(hoja.cell(row= fila, column=12).value)*1 != 0:
                                                hoja_corrugado.cell(row=articulo,column=23, value=int(1/float(hoja.cell(row= componente, column=12).value))) #UND.EMPQUE MAESTRO
                                            else:
                                                hoja_corrugado.cell(row=articulo,column=23, value=hoja.cell(row= componente, column=12).value) #UND.EMPQUE MAESTRO
                                        else:
                                            cor1 = hoja_corrugado.cell(row=articulo,column=103).value
                                            cor2 = hoja.cell(row= componente, column=8).value                            
                                            tam1 = re.findall(r'\d+',cor1)
                                            tam2 = re.findall(r'\d+',cor2)
                                            if len(tam1)>3 and len(tam2)>3:                                                
                                                if float(tam1[1])*float(tam1[2])*float(tam1[3]) > float(tam2[1])*float(tam2[2])*float(tam2[3]):
                                                    hoja_corrugado.cell(row=articulo,column=24, value=hoja.cell(row= componente, column=7).value) #CORRUGADO SECUNDARIO
                                                    hoja_corrugado.cell(row=articulo,column=99, value=(1-float(hoja.cell(row= componente, column=13).value))) #DESPERDICIO CORRUGADO SEC
                                                    if float(hoja.cell(row= componente, column=12).value)*1 != 0:
                                                        hoja_corrugado.cell(row=articulo,column=25, value=int(1/float(hoja.cell(row= componente, column=12).value))) #UND.EMPQUE SECUNDARIO
                                    else:
                                        for sp in lista_SP:
                                            if re.search(sp, hoja.cell(row= componente, column=8).value)is not None:
                                                codComponente = hoja.cell(row= componente, column=7).value
                                                for ultimo in range(1, hoja.max_row):
                                                    if codComponente == hoja.cell(row= ultimo, column=2).value and (hoja.cell(row= ultimo, column=5).value)is None:
                                                        if re.search(corrugado, hoja.cell(row= ultimo,column=8).value)is not None:
                                                            if (hoja_corrugado.cell(row=articulo,column=22).value)is None:
                                                                hoja_corrugado.cell(row=articulo,column=22, value=hoja.cell(row=ultimo, column=7).value)#CORRUGADO MASTER
                                                                break
                                                            else:
                                                                break
                                if hoja_corrugado.cell(row=articulo,column=16).value is None: #CUARTO NIVEL FORMULACION
                                    for sp in lista_SP:
                                        if re.search(sp, hoja.cell(row= componente, column=8).value)is not None:
                                            codComponente = hoja.cell(row= componente, column=7).value
                                            for ultimo in range(1, hoja.max_row):
                                                if codComponente == hoja.cell(row= ultimo, column=2).value and (hoja.cell(row= ultimo, column=5).value)is None:
                                                    for formu in formulacion:
                                                        if re.match(formu, hoja.cell(row=ultimo, column=8).value)is not None and (hoja_corrugado.cell(row=articulo,column=16).value)is None:
                                                            hoja_corrugado.cell(row=articulo,column=16, value=hoja.cell(row= ultimo, column=7).value) #FORMULACION
                if (hoja_corrugado.cell(row=articulo,column=18).value)is None:
                    codCorrugado = hoja.cell(row= fila, column=7).value
                    for componente in range(1, hoja.max_row):   #RECORRE LAS LISTAS BUSCANDO EL COD DEL SP PARA ENCONTRAR EL COLOR
                        if codCorrugado == hoja.cell(row= componente, column=2).value and (hoja.cell(row= componente, column=5).value)is None:
                            for color in colores:
                                if re.search(color, hoja.cell(row=componente, column=3).value)is not None and  (hoja_corrugado.cell(row=articulo,column=18).value)is None:
                                    hoja_corrugado.cell(row=articulo,column=18, value=color) #COLOR
                                    break
                                elif re.search(color, hoja.cell(row=componente,column=8).value)is not None and (hoja_corrugado.cell(row=articulo,column=18).value)is None:
                                    hoja_corrugado.cell(row=articulo,column=18, value=color) #COLOR
                                    break
                            if (hoja_corrugado.cell(row=articulo,column=18).value)is not None:
                                break
        for ruta in range(2, hoja_ruta.max_row):
            if codPadre == hoja_ruta.cell(row= ruta, column=3).value and (hoja_ruta.cell(row= ruta, column=5).value)is None:
                if float(hoja_ruta.cell(row= ruta, column=6).value) == 5:
                    hoja_corrugado.cell(row = articulo, column = 90, value= hoja_ruta.cell(row= ruta, column=7).value) #OP STD
                    hoja_corrugado.cell(row = articulo, column = 91, value= hoja_ruta.cell(row= ruta, column=10).value)#DESCRIPCION OP                
                elif float(hoja_ruta.cell(row= ruta, column=6).value) == 10 and re.search("AL", str(hoja_ruta.cell(row=ruta, column=14).value))is None:
                    hoja_corrugado.cell(row = articulo, column = 92, value= hoja_ruta.cell(row= ruta, column=7).value) #OP STD
                    hoja_corrugado.cell(row = articulo, column = 93, value= hoja_ruta.cell(row= ruta, column=14).value)#RECURSO
                    hoja_corrugado.cell(row = articulo, column = 94, value= hoja_ruta.cell(row= ruta, column=17).value)#INVERSO
    cod_articulos.save(nombre)
    print("TERMINA EL CORRUGADO",time.asctime(time.localtime(time.time())))

def obtener_corrugado_k42(cod_articulos):
    """TOMA EL ARCHIVO FILTRADO DE LOS ARTICULOS Y BUSCA EL CORRUGADO DE CADA ARTICULO, EN EL PROCESO SE OBTIENEN MUCHOS MAS DATOS DEL MISMO ARTICULO"""
    hoja_corrugado = cod_articulos.worksheets[0]
    libro = load_workbook('C:\Maestro\Maestro Listas k42.xlsx')
    hoja = libro.worksheets[0]
    libro_rutas = load_workbook('C:\Maestro\Maestro rutas k42.xlsx')
    hoja_ruta = libro_rutas.worksheets[0]
    lista_SP =['SP BASE','SP CONTENEDOR','SP VASO','SP TUBO','MAQUILA','SP LAM','SP BANDEJA','SP CAZUELA','SP VISOR','SP ESTUCHE','SP COPA','SP SOBRECOPA','SP ETIQUETA','SP PLIEGO']
    lista_insumo =['ETIQUETA','FUNDA','FAJA','TARJETA','TAPA MAQUILA','MATMANUF']
    formulacion = ['R ', 'PASTILLA']
    colores =['AMARILLO','AMBAR','AZUL','BEIGE','BLANCO','CAFE','CASTAÑO','DORADO','GRIS','LILA','MARRON','NARANJA','NEGRO','OPAL','PERLADO','PLATA','ROJO','ROSADO','TRANSLUCIDO','TRASLUCIDO','TRANSPARENTE','VERDE','VINOTINTO','VIOLETA']
    corrugado = "CORRUGADO"
    nombre = r"C:\Maestro\MaestroK42.xlsx" 
    for articulo in range(1, hoja_corrugado.max_row):           #RECORRE LA HOJA CON TODOS LOS PADRES DE ARTICULOS
        codPadre = hoja_corrugado.cell(row= articulo, column=1).value
        for fila in range(2, hoja.max_row):                     #RECORRE LA HOJA DE LAS LISTAS BUSCANDO EL CODIGO PADRE PARA ENCONTRAR EL SP BASE
            if codPadre == hoja.cell(row= fila, column=2).value and (hoja.cell(row= fila, column=5).value)is None: #EVALUA QUE EL COD SEA IGUAL Y PRINCIPAL
                """Primero debe buscar si tiene corrugado, si lo tiene lo trae de una
                    sino, busca el SP BASE, SP CONTENEDOR, SP VASO, SP TUBO, MAQUILA: para obtener el corrugado"""
                if re.search("TAPA", hoja.cell(row=fila,column=3).value) is not None or re.search("TUBO", hoja.cell(row=fila,column=3).value) is not None:
                    if re.search("SP TAPA", hoja.cell(row=fila,column=8).value) is not None:
                        hoja_corrugado.cell(row=articulo,column=10, value=hoja.cell(row=fila,column=7).value)#CODIGO COMPONENTE
                for color in colores:
                    if re.search(color, hoja.cell(row=fila,column=3).value)is not None and (hoja_corrugado.cell(row=articulo,column=18).value)is None:
                        hoja_corrugado.cell(row=articulo,column=18, value=color) #COLOR
                        break
                    elif re.search(color, hoja.cell(row=fila,column=8).value)is not None and (hoja_corrugado.cell(row=articulo,column=18).value)is None:
                        hoja_corrugado.cell(row=articulo,column=18, value=color) #COLOR
                        break
                for formu in formulacion:
                    if re.match(formu, hoja.cell(row= fila, column=8).value)is not None:
                        hoja_corrugado.cell(row=articulo,column=16, value=hoja.cell(row= fila, column=7).value) #FORMULACION
                for insumo in lista_insumo:                
                    if re.match(insumo, hoja.cell(row= fila, column=8).value)is not None:
                        hoja_corrugado.cell(row=articulo,column=21, value=hoja.cell(row=fila,column=7).value)#CODIGO INSUMO
                        break
                if re.search("CORRUGADO", hoja.cell(row= fila, column=8).value)is not None or re.search("BOLSA",hoja.cell(row= fila, column=8).value)is not None:
                    if re.search("BOLSA",hoja.cell(row= fila, column=8).value)is not None:
                        if (hoja_corrugado.cell(row= articulo, column=30).value) is None:
                            hoja_corrugado.cell(row=articulo,column=30, value=hoja.cell(row= fila, column=7).value) #BOLSA/ROLLO MASTER
                            hoja_corrugado.cell(row=articulo,column=104, value=hoja.cell(row= fila, column=8).value) #DESCRIPCION BOLSA MASTER
                            hoja_corrugado.cell(row=articulo,column=99, value=(1-float(hoja.cell(row= fila, column=13).value))) #DESPERDICIO BOLSA MASTER
                            if float(hoja.cell(row= fila, column=12).value)*1 != 0 :
                                hoja_corrugado.cell(row=articulo,column=31, value=int(1/float(hoja.cell(row= fila, column=12).value))) #UND.EMPQUE BOLSA MASTER
                            else:
                                hoja_corrugado.cell(row=articulo,column=31, value=hoja.cell(row= fila, column=12).value) #UND.EMPQUE BOLSA MASTER
                        else:
                            bolsa1=hoja_corrugado.cell(row= articulo, column=104).value
                            bolsa2=hoja.cell(row= fila, column=8).value
                            bol1 = bolsa1[bolsa1.find("X")-5:bolsa1.find("X")+5]                            
                            bol2 = bolsa2[bolsa2.find("X")-5:bolsa2.find("X")+5]
                            bol1 = re.findall(r'\d+',bolsa1)
                            bol2 = re.findall(r'\d+',bolsa2)
                            if len(bol1) == 2 and len(bol2)==2:
                                if float(bol1[0])*float(bol1[1]) > float(bol2[0])*float(bol2[1]):
                                    hoja_corrugado.cell(row=articulo,column=32, value=hoja.cell(row= fila, column=7).value) #BOLSA ROLLO SEC
                                    hoja_corrugado.cell(row=articulo,column=101, value=(1-float(hoja.cell(row= fila, column=13).value))) #DESPERDICIO BOLSA SEC
                                    if float(hoja.cell(row= fila, column=12).value)*1 != 0 :
                                        hoja_corrugado.cell(row=articulo,column=33, value=int(1/float(hoja.cell(row= fila, column=12).value))) #UND.EMPQUE BOLSA SECUNDARIA
                                    else:
                                        hoja_corrugado.cell(row=articulo,column=33, value=hoja.cell(row= fila, column=12).value) #UND.EMPQUE BOLSA SECUNDARIA
                                else:
                                    hoja_corrugado.cell(row=articulo,column=32, value=hoja_corrugado.cell(row=articulo,column=30).value) #BOLSA ROLLO SEC
                                    hoja_corrugado.cell(row=articulo,column=33, value=hoja_corrugado.cell(row=articulo,column=31).value) #BOLSA ROLLO SEC
                                    hoja_corrugado.cell(row=articulo,column=101, value=hoja_corrugado.cell(row=articulo,column=99).value) #DESPERDICIO BOLSA SEC
                                    hoja_corrugado.cell(row=articulo,column=30, value=hoja.cell(row= fila, column=7).value) #BOLSA ROLLO SEC
                                    hoja_corrugado.cell(row=articulo,column=99, value=(1-float(hoja.cell(row= fila, column=13).value))) #DESPERDICIO BOLSA SEC
                                    if float(hoja.cell(row= fila, column=12).value)*1 != 0 :
                                        hoja_corrugado.cell(row=articulo,column=31, value=int(1/float(hoja.cell(row= fila, column=12).value))) #UND.EMPQUE BOLSA SECUNDARIA
                                    else:
                                        hoja_corrugado.cell(row=articulo,column=31, value=hoja.cell(row= fila, column=12).value) #UND.EMPQUE BOLSA SECUNDARIA
                            else:
                                hoja_corrugado.cell(row=articulo,column=32, value="Validar descripcion")
                                hoja_corrugado.cell(row=articulo,column=33, value="Validar descripcion")                        
                    else:
                        if (hoja_corrugado.cell(row=articulo,column=22).value)is None:
                            hoja_corrugado.cell(row=articulo,column=22, value=hoja.cell(row= fila, column=7).value) #CORRUGADO MASTER
                            hoja_corrugado.cell(row=articulo,column=103, value=hoja.cell(row= fila, column=8).value) #DESCRIPCION CORRUGADO MASTER
                            hoja_corrugado.cell(row=articulo,column=98, value=(1-float(hoja.cell(row= fila, column=13).value))) #DESPERDICIO CORRUGADO MASTER
                            if float(hoja.cell(row= fila, column=12).value)*1 != 0:
                                hoja_corrugado.cell(row=articulo,column=23, value=int(1/float(hoja.cell(row= fila, column=12).value))) #UND.EMPQUE MAESTRO
                            else:
                                hoja_corrugado.cell(row=articulo,column=23, value=hoja.cell(row= fila, column=12).value) #UND.EMPQUE MAESTRO
                        else:
                            cor1 = hoja_corrugado.cell(row=articulo,column=103).value
                            cor2 = hoja.cell(row= fila, column=8).value
                            tam1 = re.findall(r'\d+',cor1)
                            tam2 = re.findall(r'\d+',cor2)
                            if len(tam1)>3 and len(tam2)>3:
                                if float(tam1[1])*float(tam1[2])*float(tam1[3]) > float(tam2[1])*float(tam2[2])*float(tam2[3]):
                                    hoja_corrugado.cell(row=articulo,column=24, value=hoja.cell(row= fila, column=7).value) #CORRUGADO SECUNDARIO
                                    hoja_corrugado.cell(row=articulo,column=99, value=(1-float(hoja.cell(row= fila, column=13).value))) #DESPERDICIO CORRUGADO SEC
                                    if float(hoja.cell(row= fila, column=12).value)*1 != 0:
                                        hoja_corrugado.cell(row=articulo,column=25, value=int(1/float(hoja.cell(row= fila, column=12).value))) #UND.EMPQUE SECUNDARIO
                                    else:
                                        hoja_corrugado.cell(row=articulo,column=25, value=hoja.cell(row= fila, column=12).value) #UND.EMPQUE SECUNDARIO
                                else:
                                    hoja_corrugado.cell(row=articulo,column=24, value=hoja_corrugado.cell(row=articulo,column=22).value) #CORRUGADO SECUNDARIO
                                    hoja_corrugado.cell(row=articulo,column=25, value=hoja_corrugado.cell(row=articulo,column=23).value) #UND EMPQUE SECUNDARIO
                                    hoja_corrugado.cell(row=articulo,column=99, value=(1-float(hoja.cell(row= fila, column=13).value))) #DESPERDICIO CORRUGADO SEC
                                    hoja_corrugado.cell(row=articulo,column=22, value=hoja.cell(row= fila, column=7).value) #CORRUGADO MASTER
                                    hoja_corrugado.cell(row=articulo,column=98, value=(1-float(hoja.cell(row= fila, column=13).value))) #DESPERDICIO CORRUGADO MASTER
                                    if float(hoja.cell(row= fila, column=12).value)*1 != 0:
                                        hoja_corrugado.cell(row=articulo,column=23, value=int(1/float(hoja.cell(row= fila, column=12).value))) #UND.EMPQUE MAESTRO
                                    else:
                                        hoja_corrugado.cell(row=articulo,column=23, value=hoja.cell(row= fila, column=12).value) #UND.EMPQUE MAESTRO                
                else:
                    for sp in lista_SP:
                        if re.search(sp, hoja.cell(row= fila, column=8).value)is not None:
                            hoja_corrugado.cell(row=articulo,column=10, value=hoja.cell(row=fila,column=7).value)#CODIGO COMPONENTE SP
                            hoja_corrugado.cell(row=articulo,column=96, value=(1-float(hoja.cell(row= fila, column=13).value))) #DESPERDICIO COMPONENTE
                            codCorrugado = hoja.cell(row= fila, column=7).value
                            for componente in range(1, hoja.max_row):   #RECORRE LAS LISTAS BUSCANDO EL COD DEL SP PARA ENCONTRAR EL COMPONENTE
                                if codCorrugado == hoja.cell(row= componente, column=2).value and (hoja.cell(row= componente, column=5).value)is None:
                                    for color in colores:
                                        if re.search(color, hoja.cell(row=componente, column=3).value)is not None and  (hoja_corrugado.cell(row=articulo,column=18).value)is None:
                                            hoja_corrugado.cell(row=articulo,column=18, value=color) #COLOR
                                            break
                                        elif re.search(color, hoja.cell(row=componente,column=8).value)is not None and (hoja_corrugado.cell(row=articulo,column=18).value)is None:
                                            hoja_corrugado.cell(row=articulo,column=18, value=color) #COLOR
                                            break
                                    des = hoja.cell(row= componente, column=8).value
                                    for formu in formulacion:
                                        if re.match(formu, hoja.cell(row=componente, column=8).value)is not None and (hoja_corrugado.cell(row=articulo,column=16).value)is None:
                                            hoja_corrugado.cell(row=articulo,column=16, value=hoja.cell(row= componente, column=7).value) #FORMULACION                                    
                                    if re.search("BOLSA",hoja.cell(row= componente, column=8).value)is not None and (hoja_corrugado.cell(row= articulo, column=32).value) is None:
                                        hoja_corrugado.cell(row=articulo,column=32, value=hoja.cell(row= componente, column=7).value) #BOLSA/ROLLO MASTER
                                        hoja_corrugado.cell(row=articulo,column=33, value=int(1/float(hoja.cell(row= componente, column=12).value))) #UND EMPQUE BOLSA MASTER
                                        hoja_corrugado.cell(row=articulo,column=100, value=(1-float(hoja.cell(row= fila, column=13).value))) #DESPERDICIO BOLSA MASTER
                                    
                                    if re.search(corrugado, des)is not None:
                                        if (hoja_corrugado.cell(row=articulo,column=22).value)is None:
                                            hoja_corrugado.cell(row=articulo,column=22, value=hoja.cell(row= componente, column=7).value) #CORRUGADO MASTER
                                            hoja_corrugado.cell(row=articulo,column=103, value=hoja.cell(row= componente, column=8).value) #DESCRIPCION CORRUGADO MASTER
                                            hoja_corrugado.cell(row=articulo,column=98, value=(1-float(hoja.cell(row= componente, column=13).value))) #DESPERDICIO CORRUGADO MASTER
                                            if float(hoja.cell(row= fila, column=12).value)*1 != 0:
                                                hoja_corrugado.cell(row=articulo,column=23, value=int(1/float(hoja.cell(row= componente, column=12).value))) #UND.EMPQUE MAESTRO
                                            else:
                                                hoja_corrugado.cell(row=articulo,column=23, value=hoja.cell(row= componente, column=12).value) #UND.EMPQUE MAESTRO
                                        else:
                                            cor1 = hoja_corrugado.cell(row=articulo,column=103).value
                                            cor2 = hoja.cell(row= componente, column=8).value                            
                                            tam1 = re.findall(r'\d+',cor1)
                                            tam2 = re.findall(r'\d+',cor2)
                                            if len(tam1)>3 and len(tam2)>3:                                                
                                                if float(tam1[1])*float(tam1[2])*float(tam1[3]) > float(tam2[1])*float(tam2[2])*float(tam2[3]):
                                                    hoja_corrugado.cell(row=articulo,column=24, value=hoja.cell(row= componente, column=7).value) #CORRUGADO SECUNDARIO
                                                    hoja_corrugado.cell(row=articulo,column=99, value=(1-float(hoja.cell(row= componente, column=13).value))) #DESPERDICIO CORRUGADO SEC
                                                    if float(hoja.cell(row= componente, column=12).value)*1 != 0:
                                                        hoja_corrugado.cell(row=articulo,column=25, value=int(1/float(hoja.cell(row= componente, column=12).value))) #UND.EMPQUE SECUNDARIO
                                    else:
                                        for sp in lista_SP:
                                            if re.search(sp, hoja.cell(row= componente, column=8).value)is not None:
                                                codComponente = hoja.cell(row= componente, column=7).value
                                                for ultimo in range(1, hoja.max_row):
                                                    if codComponente == hoja.cell(row= ultimo, column=2).value and (hoja.cell(row= ultimo, column=5).value)is None:
                                                        if re.search(corrugado, hoja.cell(row= ultimo,column=8).value)is not None:
                                                            if (hoja_corrugado.cell(row=articulo,column=22).value)is None:
                                                                hoja_corrugado.cell(row=articulo,column=22, value=hoja.cell(row=ultimo, column=7).value)#CORRUGADO MASTER
                                                                break
                                                            else:
                                                                break
                                if hoja_corrugado.cell(row=articulo,column=16).value is None: #CUARTO NIVEL FORMULACION
                                    for sp in lista_SP:
                                        if re.search(sp, hoja.cell(row= componente, column=8).value)is not None:
                                            codComponente = hoja.cell(row= componente, column=7).value
                                            for ultimo in range(1, hoja.max_row):
                                                if codComponente == hoja.cell(row= ultimo, column=2).value and (hoja.cell(row= ultimo, column=5).value)is None:
                                                    for formu in formulacion:
                                                        if re.match(formu, hoja.cell(row=ultimo, column=8).value)is not None and (hoja_corrugado.cell(row=articulo,column=16).value)is None:
                                                            hoja_corrugado.cell(row=articulo,column=16, value=hoja.cell(row= ultimo, column=7).value) #FORMULACION
                if (hoja_corrugado.cell(row=articulo,column=18).value)is None:
                    codCorrugado = hoja.cell(row= fila, column=7).value
                    for componente in range(1, hoja.max_row):   #RECORRE LAS LISTAS BUSCANDO EL COD DEL SP PARA ENCONTRAR EL COLOR
                        if codCorrugado == hoja.cell(row= componente, column=2).value and (hoja.cell(row= componente, column=5).value)is None:
                            for color in colores:
                                if re.search(color, hoja.cell(row=componente, column=3).value)is not None and  (hoja_corrugado.cell(row=articulo,column=18).value)is None:
                                    hoja_corrugado.cell(row=articulo,column=18, value=color) #COLOR
                                    break
                                elif re.search(color, hoja.cell(row=componente,column=8).value)is not None and (hoja_corrugado.cell(row=articulo,column=18).value)is None:
                                    hoja_corrugado.cell(row=articulo,column=18, value=color) #COLOR
                                    break
                            if (hoja_corrugado.cell(row=articulo,column=18).value)is not None:
                                break
        for ruta in range(2, hoja_ruta.max_row):
            if codPadre == hoja_ruta.cell(row= ruta, column=3).value and (hoja_ruta.cell(row= ruta, column=5).value)is None:
                if float(hoja_ruta.cell(row= ruta, column=6).value) == 5:
                    hoja_corrugado.cell(row = articulo, column = 90, value= hoja_ruta.cell(row= ruta, column=7).value) #OP STD
                    hoja_corrugado.cell(row = articulo, column = 91, value= hoja_ruta.cell(row= ruta, column=10).value)#DESCRIPCION OP                
                elif float(hoja_ruta.cell(row= ruta, column=6).value) == 10 and re.search("AL", str(hoja_ruta.cell(row=ruta, column=14).value))is None:
                    hoja_corrugado.cell(row = articulo, column = 92, value= hoja_ruta.cell(row= ruta, column=7).value) #OP STD
                    hoja_corrugado.cell(row = articulo, column = 93, value= hoja_ruta.cell(row= ruta, column=14).value)#RECURSO
                    hoja_corrugado.cell(row = articulo, column = 94, value= hoja_ruta.cell(row= ruta, column=17).value)#INVERSO
    cod_articulos.save(nombre)
    print("¡LISTO!")

def convertir_rutas_k40(rutas_k40):
    """TOMA EL ARCHIVO RUTAS_K40 CVS DE LAS CARPETAS COMPARTIDAS Y LO CONVIERTE DE TEXTO A COLUMNAS"""
    f = open(rutas_k40)
    csv.register_dialect('colons', delimiter='|')
    reader = csv.reader(f, dialect='colons')
    wb = Workbook()
    dest_filename = r"C:\Maestro\Maestro rutas k40.xlsx"
    ws = wb.worksheets[0]
    ws.title = "Rutas"
    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):
            column_letter = get_column_letter((column_index + 1))
            ws['%s%s'%(column_letter, (row_index + 1))].value = cell
    wb.save(filename = dest_filename)

def convertir_rutas_k42(rutas_k42):
    """TOMA EL ARCHIVO RUTAS_K42 CVS DE LAS CARPETAS COMPARTIDAS Y LO CONVIERTE DE TEXTO A COLUMNAS"""
    f = open(rutas_k42)
    csv.register_dialect('colons', delimiter='|')
    reader = csv.reader(f, dialect='colons')
    wb = Workbook()
    dest_filename = r"C:\Maestro\Maestro rutas k42.xlsx"
    ws = wb.worksheets[0]
    ws.title = "Rutas"
    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):
            column_letter = get_column_letter((column_index + 1))
            ws['%s%s'%(column_letter, (row_index + 1))].value = cell
    wb.save(filename = dest_filename)

def union(ruta):
    """TOMA LOS ARCHIVOS MAESTROS K40 Y K42 PARA UNIRLOS Y FORMAR EL MAESTRO DE ARTICULOS FINAL"""
    print("Inicia la union",time.asctime(time.localtime(time.time())))
    libro40 = load_workbook('C:\Maestro\MaestroK40.xlsx')
    libro42 = load_workbook('C:\Maestro\MaestroK42.xlsx')
    hoja40 = libro40.worksheets[0]
    hoja42 = libro42.worksheets[0]
    cont = hoja40.max_row
    cont2 = 1
    for articulo in range(2, hoja42.max_row):
        for fila in range(2, hoja40.max_row):
            hoja40.cell(row=fila, column=103, value="")
            hoja40.cell(row=fila, column=104, value="")
            if hoja42.cell(row=articulo, column=1).value == hoja40.cell(row=fila, column=1).value:
                hoja42.cell(row=articulo, column=95, value="K40/K42" )
                hoja40.cell(row=fila, column=95, value="K40/K42" )
                break            
    libro42.save(ruta)
    for articulo in range(2, hoja42.max_row):
        cont += 1
        hoja40.cell(row=cont, column=1, value=hoja42.cell(row=articulo, column=1).value)
        hoja40.cell(row=cont, column=2, value=hoja42.cell(row=articulo, column=2).value)
        hoja40.cell(row=cont, column=3, value=hoja42.cell(row=articulo, column=3).value)
        hoja40.cell(row=cont, column=4, value=hoja42.cell(row=articulo, column=4).value)
        hoja40.cell(row=cont, column=5, value=hoja42.cell(row=articulo, column=5).value)
        hoja40.cell(row=cont, column=6, value=hoja42.cell(row=articulo, column=6).value)
        hoja40.cell(row=cont, column=7, value=hoja42.cell(row=articulo, column=7).value)
        hoja40.cell(row=cont, column=8, value=hoja42.cell(row=articulo, column=8).value)
        hoja40.cell(row=cont, column=9, value=hoja42.cell(row=articulo, column=9).value)
        hoja40.cell(row=cont, column=10, value=hoja42.cell(row=articulo, column=10).value)
        hoja40.cell(row=cont, column=11, value=hoja42.cell(row=articulo, column=11).value)
        hoja40.cell(row=cont, column=15, value=hoja42.cell(row=articulo, column=15).value)
        hoja40.cell(row=cont, column=16, value=hoja42.cell(row=articulo, column=16).value)
        hoja40.cell(row=cont, column=17, value=hoja42.cell(row=articulo, column=17).value)
        hoja40.cell(row=cont, column=18, value=hoja42.cell(row=articulo, column=18).value)
        hoja40.cell(row=cont, column=19, value=hoja42.cell(row=articulo, column=19).value)
        hoja40.cell(row=cont, column=20, value=hoja42.cell(row=articulo, column=20).value)
        hoja40.cell(row=cont, column=21, value=hoja42.cell(row=articulo, column=21).value)
        hoja40.cell(row=cont, column=22, value=hoja42.cell(row=articulo, column=22).value)
        hoja40.cell(row=cont, column=23, value=hoja42.cell(row=articulo, column=23).value)
        hoja40.cell(row=cont, column=24, value=hoja42.cell(row=articulo, column=24).value)
        hoja40.cell(row=cont, column=25, value=hoja42.cell(row=articulo, column=25).value)
        hoja40.cell(row=cont, column=30, value=hoja42.cell(row=articulo, column=30).value)
        hoja40.cell(row=cont, column=31, value=hoja42.cell(row=articulo, column=31).value)
        hoja40.cell(row=cont, column=32, value=hoja42.cell(row=articulo, column=32).value)
        hoja40.cell(row=cont, column=33, value=hoja42.cell(row=articulo, column=33).value)
        hoja40.cell(row=cont, column=73, value=hoja42.cell(row=articulo, column=73).value)
        hoja40.cell(row=cont, column=74, value=hoja42.cell(row=articulo, column=74).value)
        hoja40.cell(row=cont, column=75, value=hoja42.cell(row=articulo, column=75).value)
        hoja40.cell(row=cont, column=76, value=hoja42.cell(row=articulo, column=76).value)
        hoja40.cell(row=cont, column=77, value=hoja42.cell(row=articulo, column=77).value)
        hoja40.cell(row=cont, column=78, value=hoja42.cell(row=articulo, column=78).value)
        hoja40.cell(row=cont, column=79, value=hoja42.cell(row=articulo, column=79).value)
        hoja40.cell(row=cont, column=80, value=hoja42.cell(row=articulo, column=80).value)
        hoja40.cell(row=cont, column=81, value=hoja42.cell(row=articulo, column=81).value)
        hoja40.cell(row=cont, column=82, value=hoja42.cell(row=articulo, column=82).value)
        hoja40.cell(row=cont, column=90, value=hoja42.cell(row=articulo, column=90).value)
        hoja40.cell(row=cont, column=91, value=hoja42.cell(row=articulo, column=91).value)
        hoja40.cell(row=cont, column=92, value=hoja42.cell(row=articulo, column=92).value)
        hoja40.cell(row=cont, column=93, value=hoja42.cell(row=articulo, column=93).value)
        hoja40.cell(row=cont, column=94, value=hoja42.cell(row=articulo, column=94).value)
        hoja40.cell(row=cont, column=95, value=hoja42.cell(row=articulo, column=95).value)
        hoja40.cell(row=cont, column=96, value=hoja42.cell(row=articulo, column=96).value)
        hoja40.cell(row=cont, column=97, value=hoja42.cell(row=articulo, column=97).value)
        hoja40.cell(row=cont, column=98, value=hoja42.cell(row=articulo, column=98).value)
        hoja40.cell(row=cont, column=99, value=hoja42.cell(row=articulo, column=99).value)
        hoja40.cell(row=cont, column=100, value=hoja42.cell(row=articulo, column=100).value)
        hoja40.cell(row=cont, column=101, value=hoja42.cell(row=articulo, column=101).value)
        hoja40.cell(row=cont, column=102, value=hoja42.cell(row=articulo, column=102).value)
    libro40.save(ruta)
    print("Termina la union",time.asctime(time.localtime(time.time())))
    fin = datetime.datetime.now()
    final = fin-inicio
    messagebox.showinfo(message=f"Finalizó con exito, se demoró {final}",title="Aviso")
    actualizar_registro()

def actualizar_registro():
    ruta = r'\\efiles\Temp\04_Carpak\Reportes_Oracle\HISTORIAL_MAESTRO.xlsx'
    try:
        if os.path.isfile(ruta):
            print("Ya existe")
            files = load_workbook(ruta)
            file = files.worksheets[0]
            fecha = datetime.datetime.now()
            usuario = getpass.getuser()
            file.append([fecha, usuario])
            files.save(ruta)
        else:
            files = Workbook()
            file = files.worksheets[0]
            fecha = datetime.datetime.now()
            usuario = getpass.getuser()
            file.append([fecha, usuario])
            files.save(ruta)
        messagebox.showinfo(message="El registro se actualizó",title="Actualizacion")
    except OSError as e:
        if e.erno != errno.EXXIST:
            raise
        messagebox.showinfo(message="Falló la actualizacion del registro",title="Error")
        
def hilos(articulos_k40,articulos_k42,listas_k40,listas_k42,rutas_k40,rutas_k42):
    """SE CREAN HILOS PARA EJECUTAR LAS FUNCIONES DE K40 Y K42 CASI DE FORMA PARALELA"""
    h1 = threading.Thread(target=convertir_articulos_k40(articulos_k40))
    h2 = threading.Thread(target=convertir_listas_k40(listas_k40))
    h3 = threading.Thread(target=convertir_rutas_k40(rutas_k40))
    h4 = threading.Thread(target=convertir_articulos_k42(articulos_k42))
    h5 = threading.Thread(target=convertir_listas_k42(listas_k42))
    h6 = threading.Thread(target=convertir_rutas_k42(rutas_k42))
    h1.start()
    h2.start()
    h3.start()
    h4.start()
    h5.start()
    h6.start()

def hilos_art():
    """SE CREAN HILOS PARA EJECUTAR LAS FUNCIONES DE K40 Y K42 CASI DE FORMA PARALELA"""
    cod_articulos40 = obtener_cod_articulo_k40()
    cod_articulos42 = obtener_cod_articulo_k42()
    h1 = threading.Thread(target=obtener_corrugado_k40(cod_articulos40))
    h2 = threading.Thread(target=obtener_corrugado_k42(cod_articulos42))
    h1.start()
    h2.start()

def ejecutar_archivos():
    """TOMA LAS RUTAS PREDEFINIDAS Y EJECUTA TODAS LAS FUNCIONES PARA OBTENER EL ARCHIVO FINAL"""
    global maestro, inicio
    messagebox.showinfo(message="Convirtiendo articulos, esto puede tardar",title="Convirtiendo...")
    inicio = datetime.datetime.now()
    print("Inicia el programa",time.asctime(time.localtime(time.time())))
    hilos(articulos_k40,articulos_k42,listas_k40,listas_k42,rutas_k40,rutas_k42)
    time.sleep(150)
    hilos_art()
    ruta = r"C:\Maestro\Maestro ArticulosK40K42.xlsx"
    union(ruta)
    print("Termina el programa",time.asctime(time.localtime(time.time())))
    
def ejecutar_archivos_sel(articulos_k40,articulos_k42,listas_k40,listas_k42,rutas_k40,rutas_k42):
    """TOMA LAS RUTAS DEFINIDAS POR EL USUARIO Y EJECUTA TODAS LAS FUNCIONES PARA OBTENER EL ARCHIVO FINAL"""
    global maestro, inicio
    inicio = datetime.datetime.now()
    print("Inicia el programa",time.asctime(time.localtime(time.time())))
    hilos(articulos_k40,articulos_k42,listas_k40,listas_k42,rutas_k40,rutas_k42)
    time.sleep(150)
    hilos_art()
    ruta = r"C:\Maestro\Maestro ArticulosK40K42.xlsx"
    union(ruta)
    print("Termina el programa",time.asctime(time.localtime(time.time())))
    
def guardar():
    """PREGUNTA LA RUTA DONDE SE GUARDARÁ EL ARCHIVO, TOMA LAS RUTAS PREDEFINIDAS Y EJECUTA TODAS LAS FUNCIONES PARA OBTENER EL ARCHIVO FINAL"""
    global maestro, inicio
    inicio = datetime.datetime.now()
    print("Inicia el programa",time.asctime(time.localtime(time.time())))
    maestro.iconbitmap('C:/')
    ruta = filedialog.askdirectory(parent=maestro)
    print("Se guardará en esta ruta:", ruta)
    if len(ruta)>0 :
        messagebox.showinfo(message="Se guardó con exito",title="Aviso")
        hilos(articulos_k40,articulos_k42,listas_k40,listas_k42,rutas_k40,rutas_k42)
        time.sleep(150)
        hilos_art()
        ruta = ruta + "\Maestro ArticulosK40K42.xlsx"
        union(ruta)

def cargar_archivos():
    """TOMA LAS RUTAS DEFINIDAS POR EL USUARIO Y EJECUTA TODAS LAS FUNCIONES PARA OBTENER EL ARCHIVO FINAL"""
    global maestro
    maestro.iconbitmap('C:/')
    articulos_k40 = None
    articulos_k42 = None
    listas_k40 = None
    listas_k42 = None
    rutas_k40 = None
    rutas_k42 = None
    files = filedialog.askopenfilenames(parent=maestro,initialdir="/Escritorio", filetypes=[("all files","*.*")])
    for i in files:
        print (i)
        if re.search('LISTAS_K40',i):
            listas_k40 = "r'"+i
        elif re.search('LISTAS_K42',i):
            listas_k42 = "r'"+i
        elif re.search('ARTICULOS_K40',i):
            articulos_k40 = "r'"+i
        elif re.search('ARTICULOS_K42',i):
            articulos_k42 = "r'"+i
        elif re.search('RUTAS_K40',i):
            rutas_k40 = "r'"+i
        elif re.search('RUTAS_K42',i):
            rutas_k42 = "r'"+i
    if listas_k40 is None or listas_k42 is None or articulos_k40 is None or articulos_k42 is None or rutas_k40 is None or rutas_k42 is None:
        messagebox.showinfo(message="Faltan archivos, por favor vuelve a intentarlo",title="Aviso")
    else:
        ejecutar_archivos_sel(articulos_k40,articulos_k42,listas_k40,listas_k42,rutas_k40,rutas_k42)
        messagebox.showinfo(message="Convirtiendo articulos, esto puede tardar",title="Convirtiendo...")

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def menu():
    """FUNCION MAIN LA CUAL SE ENCARGA DE MOSTRAR LA PARTE VISUAL Y PODES EJECUTAR LAS FUNCIONES DEL PROGRAMA"""
    global maestro, insertar, inserta,links
    maestro = Tk()
    maestro.title("MAESTRO DE ARTICULOS")
    maestro.resizable(0, 0)
    path = resource_path("carvajal.ico")
    maestro.iconbitmap(path)
    links = []
    m = Canvas(maestro, width=455, height = 300, bg ="gray26")
    m.pack(fill=X)
    path = resource_path("descripcion.png")
    fondo = PhotoImage(file=path)
    m.create_image(230,152,image=fondo, anchor="center")
    try:
        boton_guardar = Button(maestro,text="Guardar en:",font="Arial",width=14,command=guardar,bg="deep sky blue").place(x=10, y=18)
        boton_ejecutar = Button(maestro,text="Ejecutar archivos",font="Arial",width=14,command=ejecutar_archivos,bg="yellow green").place(x=10, y=95)
        boton_buscar = Button(maestro, text="Cargar archivos", font="Arial",width=14, command=cargar_archivos,bg="dark violet").place(x=10, y=171)
        boton_salir = Button(maestro, text="Salir", font="Arial",width=14,command=maestro.destroy, bg="dark orange").place(x=10, y=249)
    except Exception as e:
        messagebox.showinfo(message="No se pueden ejecutar los archivos por alguna de estas razones: \n-Acceso a ruta denegado.\n-Coneccion a red fallida.\n-Faltan archivos para su ejecucion.",title="Error")
        print("ERROR : "+str(e))  
    maestro.mainloop()
menu()
